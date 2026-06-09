"""
报告导出 API 路由
生成 Excel/CSV 格式的成绩单、考试报告、点名记录、课堂互动数据等
"""
import csv
import io
import json
from datetime import datetime

from fastapi import APIRouter, HTTPException, Request, Query
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from backend.api.dependencies import get_current_user
from backend.database import execute_query
from backend.question_db import (
    execute_query as q_execute_query,
    execute_query_one,
)
from backend.logger import logger

router = APIRouter()

# ── 样式常量 ──
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1677FF", end_color="1677FF", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(horizontal="center", vertical="center")
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
TITLE_FONT = Font(name="微软雅黑", bold=True, size=14)


def _style_header(ws, row: int, cols: int):
    """给表头行设置样式"""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = BORDER


def _style_cells(ws, start_row: int, end_row: int, cols: int):
    """给数据区域设置样式"""
    for row in range(start_row, end_row + 1):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = CELL_ALIGNMENT
            cell.border = BORDER


def _auto_width(ws, cols: int, max_width: int = 40):
    """自动调整列宽"""
    for col in range(1, cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, max_width)


def _excel_response(wb, filename: str) -> StreamingResponse:
    """将 workbook 转为 StreamingResponse"""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # 使用 RFC 5987 编码支持中文文件名
    from urllib.parse import quote
    encoded_name = quote(filename, safe='')
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}",
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


# ── 1. 导出课堂积分 ──

@router.get("/scores", summary="导出课堂积分表 (Excel)")
async def export_scores(
    request: Request,
    teacher: str = Query("", description="教师用户名"),
    grade: str = Query("", description="年级"),
    cls: str = Query("", description="班级"),
):
    """导出班级积分表为 Excel"""
    user = get_current_user(request)
    username = user["username"]
    role = user.get("role", 2)

    if role not in (0, 1):
        raise HTTPException(status_code=403, detail="权限不足")

    # 确定查询的教师
    query_teacher = teacher or username

    # 构建条件
    conditions = ["teacher_username = ?"]
    params = [query_teacher]
    if grade:
        conditions.append("grade = ?")
        params.append(grade)
    if cls:
        conditions.append("class_name = ?")
        params.append(cls)

    where = " AND ".join(conditions)
    rows = execute_query(
        f"""SELECT grade, class_name, student_name, score, updated_at
            FROM scores WHERE {where}
            ORDER BY grade, class_name, score DESC""",
        tuple(params),
    )

    if not rows:
        raise HTTPException(status_code=404, detail="没有找到积分数据")

    # 计算排名
    ranked = []
    seen_classes = set()
    rank_in_class = {}
    for r in rows:
        key = f"{r[0]}|{r[1]}"
        if key not in seen_classes:
            seen_classes.add(key)
            rank_in_class[key] = 1
        ranked.append({
            "grade": r[0],
            "class": r[1],
            "name": r[2],
            "score": r[3],
            "updated_at": r[4] or "",
            "rank": rank_in_class[key],
        })
        rank_in_class[key] += 1

    # 创建 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "课堂积分"

    # 标题
    title = f"课堂积分表"
    if grade:
        title += f" - {grade}"
    if cls:
        title += f" - {cls}班"
    ws.merge_cells("A1:E1")
    ws.cell(1, 1, title).font = TITLE_FONT

    headers = ["年级", "班级", "学生姓名", "积分", "排名", "最后更新"]
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    _style_header(ws, 3, len(headers))

    for idx, r in enumerate(ranked):
        row = idx + 4
        ws.cell(row, 1, r["grade"])
        ws.cell(row, 2, r["class"])
        ws.cell(row, 3, r["name"])
        ws.cell(row, 4, r["score"])
        ws.cell(row, 5, r["rank"])
        ws.cell(row, 6, r["updated_at"])

    _style_cells(ws, 4, len(ranked) + 3, len(headers))
    _auto_width(ws, len(headers))

    filename = f"课堂积分_{grade or '全部'}_{cls or '全部'}.xlsx"
    return _excel_response(wb, filename)


# ── 2. 导出考试结果 ──

@router.get("/exam/{exam_id}", summary="导出考试成绩报告 (Excel)")
async def export_exam_result(exam_id: int, request: Request):
    """导出指定考试的完整成绩报告为 Excel"""
    user = get_current_user(request)
    username = user["username"]
    role = user.get("role", 2)

    if role == 2:
        raise HTTPException(status_code=403, detail="权限不足")

    exam = execute_query_one("SELECT * FROM exams WHERE id = ?", (exam_id,))
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")

    # 获取提交记录（exam_attempts 在 questions.db 中）
    attempts = q_execute_query(
        """SELECT * FROM exam_attempts
           WHERE exam_id = ? AND status IN ('submitted', 'graded')
           ORDER BY score DESC""",
        (exam_id,),
    )

    # 批量补充学生姓名（避免 N+1 查询）
    if attempts:
        usernames = [a['student_username'] for a in attempts]
        placeholders = ",".join("?" for _ in usernames)
        user_rows = execute_query(
            f"SELECT username, name FROM users WHERE username IN ({placeholders})",
            tuple(usernames),
        )
        name_map = {r[0]: r[1] for r in user_rows} if user_rows else {}
        for a in attempts:
            a['student_real_name'] = name_map.get(a['student_username'], a.get('student_name', ''))

    # 获取考试的题目（exam_questions 和 question_bank 在 questions.db 中）
    questions = q_execute_query(
        """SELECT eq.id as eq_id, eq.score as question_score, qb.id as qid,
                  qb.type, qb.question_text, qb.correct_answer, qb.options
           FROM exam_questions eq
           JOIN question_bank qb ON eq.question_id = qb.id
           WHERE eq.exam_id = ?
           ORDER BY eq.sort_order""",
        (exam_id,),
    )

    wb = openpyxl.Workbook()

    # ── Sheet 1: 成绩总表 ──
    ws1 = wb.active
    ws1.title = "成绩总表"

    title = f"考试报告 - {exam['title']}"
    ws1.merge_cells("A1:H1")
    ws1.cell(1, 1, title).font = TITLE_FONT

    # 考试信息
    info = [
        ("科目", exam["subject"]),
        ("总分", str(exam["total_score"])),
        ("及格分", str(exam["pass_score"])),
        ("考试时长", f"{exam['duration']} 分钟"),
        ("状态", {"draft": "草稿", "published": "已发布", "ended": "已结束"}.get(exam["status"], exam["status"])),
    ]
    for i, (k, v) in enumerate(info):
        ws1.cell(3, i * 2 + 1, k).font = Font(bold=True)
        ws1.cell(3, i * 2 + 2, v)

    # 统计数据
    scores_list = [a["score"] for a in attempts]
    total = len(scores_list)
    avg_score = round(sum(scores_list) / max(total, 1), 1) if scores_list else 0
    pass_count = sum(1 for s in scores_list if s >= exam["pass_score"])
    max_score = max(scores_list) if scores_list else 0
    min_score = min(scores_list) if scores_list else 0

    stats = [
        ("参考人数", total),
        ("平均分", avg_score),
        ("最高分", max_score),
        ("最低分", min_score),
        ("及格人数", pass_count),
        ("及格率", f"{round(pass_count / max(total, 1) * 100, 1)}%"),
    ]
    for i, (k, v) in enumerate(stats):
        ws1.cell(4, i * 2 + 1, k).font = Font(bold=True)
        ws1.cell(4, i * 2 + 2, v)

    # 表头
    headers1 = ["排名", "学生姓名", "用户名", "得分", "总分", "是否及格", "提交时间"]
    row_start = 6
    for i, h in enumerate(headers1, 1):
        ws1.cell(row_start, i, h)
    _style_header(ws1, row_start, len(headers1))

    for idx, a in enumerate(attempts):
        row = row_start + 1 + idx
        ws1.cell(row, 1, idx + 1)
        ws1.cell(row, 2, a.get("student_real_name") or a.get("student_name") or "")
        ws1.cell(row, 3, a["student_username"])
        ws1.cell(row, 4, a["score"])
        ws1.cell(row, 5, exam["total_score"])
        passed = "✅ 及格" if a["score"] >= exam["pass_score"] else "❌ 未及格"
        ws1.cell(row, 6, passed)
        ws1.cell(row, 7, a.get("submitted_at") or "")

    _style_cells(ws1, row_start + 1, row_start + len(attempts), len(headers1))
    _auto_width(ws1, len(headers1))

    # ── Sheet 2: 逐题分析 ──
    if questions:
        ws2 = wb.create_sheet("逐题分析")
        ws2.merge_cells("A1:F1")
        ws2.cell(1, 1, f"逐题分析 - {exam['title']}").font = TITLE_FONT

        headers2 = ["题号", "题型", "题目内容", "正确答案", "分值", "正确率"]
        for i, h in enumerate(headers2, 1):
            ws2.cell(3, i, h)
        _style_header(ws2, 3, len(headers2))

        type_map = {"single": "单选", "multiple": "多选", "true_false": "判断", "short": "简答",
                     "fill": "填空", "essay": "作文", "subjective": "主观题"}

        for idx, q in enumerate(questions):
            row = 4 + idx
            ws2.cell(row, 1, idx + 1)
            ws2.cell(row, 2, type_map.get(q["type"], q["type"]))
            # 截取题目内容
            q_text = q["question_text"]
            if len(q_text) > 80:
                q_text = q_text[:77] + "..."
            ws2.cell(row, 3, q_text)
            ws2.cell(row, 4, q["correct_answer"])
            ws2.cell(row, 5, q["question_score"])

            # 计算正确率
            correct_count = 0
            for a in attempts:
                if a.get("answers"):
                    try:
                        answers = json.loads(a["answers"]) if isinstance(a["answers"], str) else a["answers"]
                    except (json.JSONDecodeError, TypeError):
                        answers = {}
                else:
                    answers = {}
                # 题目答案的 key 在 answers 字典中
                q_key = str(q["qid"])
                student_ans = answers.get(q_key) or answers.get(str(q["eq_id"])) or ""
                if student_ans and student_ans == q["correct_answer"]:
                    correct_count += 1

            rate = round(correct_count / max(len(attempts), 1) * 100, 1)
            ws2.cell(row, 6, f"{rate}%")

        _style_cells(ws2, 4, 4 + len(questions) - 1, len(headers2))
        _auto_width(ws2, len(headers2), max_width=50)

    filename = f"考试报告_{exam['title']}.xlsx"
    return _excel_response(wb, filename)


# ── 3. 导出点名记录 ──

@router.get("/rollcall", summary="导出点名记录 (Excel)")
async def export_rollcall(
    request: Request,
    teacher: str = Query("", description="教师用户名"),
    grade: str = Query("", description="年级"),
    cls: str = Query("", description="班级"),
    start_date: str = Query("", description="开始日期 YYYY-MM-DD"),
    end_date: str = Query("", description="结束日期 YYYY-MM-DD"),
):
    """导出点名历史记录为 Excel"""
    user = get_current_user(request)
    username = user["username"]
    role = user.get("role", 2)

    if role not in (0, 1):
        raise HTTPException(status_code=403, detail="权限不足")

    conditions = []
    params: list = []

    if role == 1:
        conditions.append("teacher_username = ?")
        params.append(username)
    elif teacher:
        conditions.append("teacher_username = ?")
        params.append(teacher)

    if grade:
        conditions.append("grade = ?")
        params.append(grade)
    if cls:
        conditions.append("class_name = ?")
        params.append(cls)
    if start_date:
        conditions.append("created_at >= ?")
        params.append(start_date)
    if end_date:
        conditions.append("created_at <= ?")
        params.append(end_date + " 23:59:59")

    where = " AND ".join(conditions) if conditions else "1=1"
    rows = execute_query(
        f"""SELECT teacher_username, grade, class_name, student_name, result, points, created_at
            FROM rollcall_history WHERE {where}
            ORDER BY created_at DESC""",
        tuple(params),
    )

    if not rows:
        raise HTTPException(status_code=404, detail="没有找到点名记录")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "点名记录"

    title = "点名记录"
    if grade:
        title += f" - {grade}"
    if cls:
        title += f" - {cls}班"
    ws.merge_cells("A1:G1")
    ws.cell(1, 1, title).font = TITLE_FONT

    headers = ["教师", "年级", "班级", "学生姓名", "回答结果", "积分变动", "时间"]
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    _style_header(ws, 3, len(headers))

    result_map = {"1": "正确", "0": "错误", "": "待定"}
    for idx, r in enumerate(rows):
        row = 4 + idx
        ws.cell(row, 1, r[0])
        ws.cell(row, 2, r[1])
        ws.cell(row, 3, r[2])
        ws.cell(row, 4, r[3])
        ws.cell(row, 5, result_map.get(str(r[4] or ""), str(r[4] or "待定")))
        ws.cell(row, 6, r[5])
        ws.cell(row, 7, r[6] or "")

    _style_cells(ws, 4, len(rows) + 3, len(headers))
    _auto_width(ws, len(headers))

    filename = f"点名记录_{grade or '全部'}_{cls or '全部'}.xlsx"
    return _excel_response(wb, filename)


# ── 4. 导出任务提交记录 ──

@router.get("/tasks", summary="导出任务提交记录 (Excel)")
async def export_tasks(
    request: Request,
    task_id: str = Query("", description="任务ID"),
):
    """导出任务提交记录为 Excel"""
    user = get_current_user(request)
    username = user["username"]
    role = user.get("role", 2)

    if role not in (0, 1):
        raise HTTPException(status_code=403, detail="权限不足")

    # 获取任务列表
    if task_id:
        tasks = execute_query(
            "SELECT * FROM tasks WHERE id = ?",
            (task_id,),
        )
    elif role == 0:
        tasks = execute_query(
            "SELECT * FROM tasks ORDER BY created_at DESC",
        )
    else:
        tasks = execute_query(
            "SELECT * FROM tasks WHERE creator_username = ? ORDER BY created_at DESC",
            (username,),
        )

    if not tasks:
        raise HTTPException(status_code=404, detail="没有找到任务")

    wb = openpyxl.Workbook()
    first_sheet = True

    for t in tasks:
        if first_sheet:
            ws = wb.active
            ws.title = _safe_sheet_name(t[2][:20])  # task name as sheet name
            first_sheet = False
        else:
            ws = wb.create_sheet(title=_safe_sheet_name(t[2][:20]))

        # 标题
        ws.merge_cells("A1:E1")
        ws.cell(1, 1, f"任务提交记录 - {t[2]}").font = TITLE_FONT
        ws.cell(2, 1, f"描述: {t[3] or ''}")
        ws.cell(2, 2, f"状态: {t[4]}")
        ws.cell(2, 4, f"创建时间: {t[5] or ''}")

        # 获取提交记录
        task_id_val = t[0]
        submissions = execute_query(
            """SELECT ts.submitted_at, ts.student_username, u.name
               FROM task_submissions ts
               LEFT JOIN users u ON ts.student_username = u.username
               WHERE ts.task_id = ?
               ORDER BY ts.submitted_at DESC""",
            (task_id_val,),
        )

        headers = ["序号", "学生用户名", "学生姓名", "提交时间"]
        for i, h in enumerate(headers, 1):
            ws.cell(4, i, h)
        _style_header(ws, 4, len(headers))

        for idx, s in enumerate(submissions):
            row = 5 + idx
            ws.cell(row, 1, idx + 1)
            ws.cell(row, 2, s[1])
            ws.cell(row, 3, s[2] or "")
            ws.cell(row, 4, s[0] or "")

        _style_cells(ws, 5, 5 + len(submissions) - 1, len(headers))
        _auto_width(ws, len(headers))

    filename = f"任务提交记录.xlsx"
    return _excel_response(wb, filename)


# ── 5. 导出学情进度 ──

@router.get("/progress", summary="导出学情进度 (Excel)")
async def export_progress(
    request: Request,
    course_id: int = Query(None, description="课程 ID"),
    grade: str = Query(None, description="年级"),
    class_name: str = Query(None, description="班级"),
):
    """导出学情进度（知识点完成情况）为 Excel"""
    user = get_current_user(request)
    role = user.get("role", 2)
    if role not in (0, 1):
        raise HTTPException(status_code=403, detail="权限不足")

    from backend.api.curriculum_router import get_class_progress_overview as fetch_progress

    # 复用进度总览接口的数据逻辑
    raw = await fetch_progress(request, course_id, grade, class_name)
    students = raw.get("students", [])

    if not students:
        raise HTTPException(status_code=404, detail="没有找到进度数据")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学情进度"

    title = "学情进度报告"
    if grade:
        title += f" - {grade}"
    if class_name:
        title += f" - {class_name}班"
    ws.merge_cells("A1:H1")
    ws.cell(1, 1, title).font = TITLE_FONT

    headers = ["姓名", "年级", "班级", "课程", "知识点总数", "已完成", "完成率", "状态"]
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    _style_header(ws, 3, len(headers))

    row_idx = 4
    for stu in students:
        stu_courses = stu.get("courses") or []
        if not stu_courses:
            ws.cell(row_idx, 1, stu.get("name", ""))
            ws.cell(row_idx, 2, stu.get("grade", ""))
            ws.cell(row_idx, 3, stu.get("class", ""))
            ws.cell(row_idx, 4, "—")
            ws.cell(row_idx, 5, 0)
            ws.cell(row_idx, 6, 0)
            ws.cell(row_idx, 7, "0%")
            ws.cell(row_idx, 8, "无课程")
            row_idx += 1
        else:
            for c in stu_courses:
                ws.cell(row_idx, 1, stu.get("name", ""))
                ws.cell(row_idx, 2, stu.get("grade", ""))
                ws.cell(row_idx, 3, stu.get("class", ""))
                ws.cell(row_idx, 4, c.get("course_name", ""))
                ws.cell(row_idx, 5, c.get("total_kps", 0))
                ws.cell(row_idx, 6, c.get("completed_kps", 0))
                rate = c.get("rate", 0)
                ws.cell(row_idx, 7, f"{rate}%")
                status = "✅ 已完成" if rate >= 100 else ("🔄 进行中" if rate > 0 else "⏳ 未开始")
                ws.cell(row_idx, 8, status)
                row_idx += 1

    _style_cells(ws, 4, row_idx - 1, len(headers))
    _auto_width(ws, len(headers))

    filename = f"学情进度_{grade or '全部'}_{class_name or '全部'}.xlsx"
    return _excel_response(wb, filename)


def _safe_sheet_name(name: str) -> str:
    """Excel sheet name max 31 chars, no special chars"""
    safe = "".join(c if c.isalnum() or c in "_ -" else "_" for c in name)
    return safe[:31] or "Sheet"


# ── CSV 响应辅助 ──

def _csv_response(rows: list[list], headers: list[str], filename: str) -> StreamingResponse:
    """将数据转为 CSV 响应（带 UTF-8 BOM，Excel 可直接打开）"""
    buf = io.StringIO()
    # 写入 BOM + 内容
    import codecs
    bom = codecs.BOM_UTF8.decode('utf-8')
    buf.write(bom)
    writer = csv.writer(buf)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    buf.seek(0)
    from urllib.parse import quote
    encoded_name = quote(filename, safe='')
    return StreamingResponse(
        iter([buf.getvalue().encode('utf-8-sig')]),
        media_type="text/csv; charset=utf-8-sig",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}",
        },
    )


# ── 6. 导出随堂测验结果 ──

@router.get("/quiz/{quiz_id}", summary="导出随堂测验结果 (CSV)")
async def export_quiz_result(quiz_id: int, request: Request,
                              token: str = Query("", description="JWT token 用于 window.open 下载")):
    """导出随堂测验的答题结果为 CSV 或 Excel"""
    # 支持 token 参数认证（用于 window.open 下载）
    if token:
        request.state.user = None
        from backend.auth import decode_jwt_token
        payload = decode_jwt_token(token)
        if payload:
            request.state.user = payload
    user = get_current_user(request)
    role = user.get("role", 2)
    if role not in (0, 1):
        raise HTTPException(status_code=403, detail="权限不足")

    # 获取测验信息
    quiz = execute_query("SELECT * FROM interaction_quizzes WHERE id = ?", (quiz_id,))
    if not quiz:
        raise HTTPException(status_code=404, detail="测验不存在")
    quiz = quiz[0]
    questions = json.loads(quiz[4]) if isinstance(quiz[4], str) else quiz[4]

    # 获取答题记录
    answers = execute_query(
        """SELECT student_username, answers, score, submitted_at
           FROM interaction_quiz_answers WHERE quiz_id = ? ORDER BY score DESC""",
        (quiz_id,),
    )

    csv_headers = ["学生", "得分"]
    for i, q in enumerate(questions):
        csv_headers.append(f"第{i + 1}题答案")
        csv_headers.append(f"第{i + 1}题是否正确")

    csv_rows = []
    for a in answers:
        student = a[0]
        ans_data = json.loads(a[1]) if isinstance(a[1], str) else a[1]
        score = a[2]
        row = [student, str(score)]
        for i, q in enumerate(questions):
            user_ans = ""
            correct = ""
            for item in ans_data:
                if item.get("question_index") == i:
                    user_ans = item.get("answer", "")
                    correct_ans = q.get("answer", "")
                    if isinstance(q.get("options"), dict):
                        correct_ans = q.get("answer", "")
                    is_correct = "是" if str(user_ans).upper() == str(correct_ans).upper() else "否"
                    row.append(user_ans)
                    row.append(is_correct)
                    break
            else:
                row.append("")
                row.append("")
        csv_rows.append(row)

    filename = f"随堂测验_{quiz[2]}.csv"
    return _csv_response(csv_rows, csv_headers, filename)


# ── 7. 导出投票结果 ──

@router.get("/poll/{poll_id}", summary="导出投票结果 (CSV)")
async def export_poll_result(poll_id: int, request: Request,
                              token: str = Query("", description="JWT token 用于 window.open 下载")):
    """导出投票结果为 CSV"""
    # 支持 token 参数认证（用于 window.open 下载）
    if token:
        request.state.user = None
        from backend.auth import decode_jwt_token
        payload = decode_jwt_token(token)
        if payload:
            request.state.user = payload
    user = get_current_user(request)
    role = user.get("role", 2)
    if role not in (0, 1):
        raise HTTPException(status_code=403, detail="权限不足")

    poll = execute_query("SELECT * FROM interaction_polls WHERE id = ?", (poll_id,))
    if not poll:
        raise HTTPException(status_code=404, detail="投票不存在")
    poll = poll[0]
    options = json.loads(poll[3]) if isinstance(poll[3], str) else poll[3]

    # 获取投票记录
    votes = execute_query(
        """SELECT student_username, selected_option FROM interaction_poll_votes
           WHERE poll_id = ? ORDER BY student_username""",
        (poll_id,),
    )

    csv_headers = ["学生", "所选选项编号", "所选选项内容"]
    csv_rows = []
    for v in votes:
        student = v[0]
        opt_idx = v[1]
        opt_text = options[opt_idx] if opt_idx < len(options) else f"选项{opt_idx}"
        csv_rows.append([student, str(opt_idx), opt_text])

    filename = f"投票结果_{poll[2]}.csv"
    return _csv_response(csv_rows, csv_headers, filename)
